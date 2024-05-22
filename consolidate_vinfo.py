from openpyxl import Workbook, load_workbook
import os
from tkinter import Tk, filedialog

def select_files():
    root = Tk()
    root.withdraw()  # Hide the root window
    file_paths = filedialog.askopenfilenames(
        title="Select Excel Files",
        filetypes=[("Excel files", "*.xlsx"), ("Excel Macro-Enabled", "*.xlsm"), ("Excel Templates", "*.xltx"), ("Macro-Enabled Templates", "*.xltm")]
    )
    root.destroy()  # Properly close the Tkinter root window
    return file_paths

def consolidate_workbooks(file_paths):
    output_file = 'consolidated_vInfo.xlsx'
    columns_of_interest = ["VM", "Powerstate", "Template", "CPUs", "Memory", 
                            "Provisioned MiB", "In Use MiB", "Unshared MiB", "Datacenter", 
                            "Cluster", "Host", "OS according to the configuration file", 
                            "OS according to the VMware Tools"]
    if not os.path.exists(output_file):
        wb = Workbook()
        ws = wb.active
        ws.append(columns_of_interest)
        wb.save(output_file)

    out_wb = load_workbook(output_file)
    out_sheet = out_wb.active

    for file_path in file_paths:
        src_wb = load_workbook(file_path, data_only=True)
        if "vInfo" in src_wb.sheetnames:
            src_sheet = src_wb["vInfo"]
            header = {cell.value: idx for idx, cell in enumerate(src_sheet[1]) if cell.value in columns_of_interest}
            max_col_index = max(header.values(), default=0) + 1  # Ensure we get all columns up to the maximum needed
            
            for row in src_sheet.iter_rows(min_row=2, max_col=max_col_index, values_only=True):
                out_row = [(row[header[col]] if col in header and len(row) > header[col] else None) for col in columns_of_interest]
                out_sheet.append(out_row)
    
    out_wb.save(output_file)
    out_wb.close()

if __name__ == "__main__":
    print("Select the Excel files to consolidate:")
    files = select_files()
    if files:
        consolidate_workbooks(files)
        print("Consolidation complete. Saved as 'consolidated_vInfo.xlsx'")
    else:
        print("No files selected.")
